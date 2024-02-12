import openpyxl

class GerenciadorProdutos:
    def __init__(self, workbook):
        self.book = workbook
        self.products_page = self.book.active
        self.products_page.title = 'Produtos'
        self.products_page.append(["Produto", "Quantidade", "Preço"])

    def inserir_produto(self):
        nome = input("Informe o nome do Produto: ")
        preco = float(input("Informe o preço do Produto: "))
        quantidade = int(input("Informe a quantidade do Produto: "))
        
        self.products_page.append([nome, preco, quantidade])
        self.book.save("app/Planilha_Produtos.xlsx")
        print("Produto Inserido com Sucesso")

    def alterar_produto(self):
        print("Produtos:")
        for row in self.products_page.iter_rows(min_row=2, max_row=self.products_page.max_row):
            print(row[0].value)
        
        produto = input("Digite o nome do produto que deseja alterar: ")
        for row in self.products_page.iter_rows(min_row=2, max_row=self.products_page.max_row):
            if row[0].value == produto:
                quantidade = input("Digite a nova quantidade: ")
                preco = input("Digite o novo preço: ")
                row[1].value = quantidade
                row[2].value = preco
                self.book.save("app/Planilha_Produtos.xlsx")
                print("Produto alterado com sucesso!")
                return
        print("Produto não encontrado.")

    def visualizar_produtos(self):
        print("Produtos:")
        for row in self.products_page.iter_rows(min_row=2, max_row=self.products_page.max_row):
            print(f"Produto: {row[0].value}, Quantidade: {row[1].value}, Preço: {row[2].value}")


book = openpyxl.load_workbook("app/Planilha_Produtos.xlsx")
gerenciador = GerenciadorProdutos(book)

while True:
    print("\nMenu:")
    print("1. Inserir produto")
    print("2. Alterar produto")
    print("3. Visualizar produtos")
    print("4. Sair")

    escolha = input("Escolha uma opção: ")

    if escolha == "1":
        gerenciador.inserir_produto()
    elif escolha == "2":
        gerenciador.alterar_produto()
    elif escolha == "3":
        gerenciador.visualizar_produtos()
    elif escolha == "4":
        print("Saindo...")
        break
    else:
        print("Opção inválida. Tente novamente.")
